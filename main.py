#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import asyncio
import json
import os
import random
from datetime import datetime
from pathlib import Path

import feedparser
import groq
import openpyxl
import pytz
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from telegram import Bot
from tenacity import retry, stop_after_attempt, wait_exponential

# ========== ENV ==========
GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID")
TIMEZONE = pytz.timezone("Asia/Dhaka")

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
EXCEL_FILE = DATA_DIR / "Vocabulary_Bank.xlsx"
HISTORY_FILE = "history.json"

RSS_FEEDS = {
    "The Daily Star": "https://www.thedailystar.net/rss.xml",
    "Dhaka Tribune": "https://www.dhakatribune.com/feed",
    "The Business Standard": "https://www.tbsnews.net/rss.xml",
}

# --------------------------------------------------------------
# HELPERS
# --------------------------------------------------------------

def safe_excel_value(value):
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return str(value)
    return str(value)

def fetch_headlines():
    all_entries = []
    for source, url in RSS_FEEDS.items():
        try:
            feed = feedparser.parse(url)
            for entry in feed.entries[:5]:
                all_entries.append({"headline": entry.title, "source": source})
        except:
            pass
    seen = set()
    unique = []
    for entry in all_entries:
        if entry["headline"] not in seen:
            seen.add(entry["headline"])
            unique.append(entry)
    return unique[:25]

def load_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, "r") as f:
            return json.load(f).get("words", [])
    return []

def save_history(words):
    with open(HISTORY_FILE, "w") as f:
        json.dump({"words": words}, f)

# --------------------------------------------------------------
# AI CALLS
# --------------------------------------------------------------

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10))
def generate_vocab_and_summary(client, headlines_data, past_words):
    headlines_text = "\n".join([f"- [{entry['source']}] {entry['headline']}" for entry in headlines_data])
    exclude = ", ".join(past_words[-50:])

    prompt = f"""
You are an expert BCS and Bank job exam mentor.

Today's headlines from Bangladeshi newspapers (with sources):
{headlines_text}

### Task 1: Generate 100 vocabulary words (numbered 1-100) from these headlines.
- Difficulty: 30 Basic, 40 Intermediate, 30 Advanced.
- Do NOT repeat: {exclude}
- For each word provide EXACTLY these keys:
  "sl", "word", "pos", "level", "bengali", "definition", "synonyms", "antonyms", "example", "category"
- The synonyms and antonyms must be given as a comma-separated string (not a list).

### Task 2: Write a detailed 5-7 bullet Bengali summary (in Bengali script) of the most important topics covered in the headlines. Include the newspaper names. The summary must be at least 100 characters long.

Return JSON with keys: "vocab_list" and "bengali_summary".
"""
    response = client.chat.completions.create(
        messages=[
            {"role": "system", "content": "You are an expert Bengali lexicographer. Always output valid JSON. For synonyms and antonyms, use a comma-separated string, not a list."},
            {"role": "user", "content": prompt}
        ],
        model="llama-3.3-70b-versatile",
        temperature=0.3,
        response_format={"type": "json_object"},
    )
    data = json.loads(response.choices[0].message.content)
    vocab = data.get("vocab_list", [])
    summary = data.get("bengali_summary", "")
    if not summary or len(summary.strip()) < 20:
        summary = build_fallback_summary(headlines_data)
    return vocab, summary

def build_fallback_summary(headlines_data):
    sources = {}
    for entry in headlines_data:
        src = entry["source"]
        if src not in sources:
            sources[src] = []
        sources[src].append(entry["headline"])

    summary = "📰 আজকের প্রধান সংবাদ (সূত্র সহ)\n\n"
    for src, headlines in sources.items():
        summary += f"▪️ {src}\n"
        for h in headlines[:3]:
            summary += f"   - {h}\n"
        summary += "\n"
    summary += "🔹 অন্যান্য সংবাদপত্র থেকেও গুরুত্বপূর্ণ শব্দ সংগ্রহ করা হয়েছে।"
    return summary

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10))
def generate_mcqs(client, vocab_list):
    selected = random.sample(vocab_list, min(10, len(vocab_list)))
    word_list = [w["word"] for w in selected]

    prompt = f"""
Generate 10 multiple-choice questions (MCQs) from these vocabulary words: {', '.join(word_list)}.
For each question, provide:
- "question": the question text
- "options": a list of 4 options (A, B, C, D)
- "answer": the correct option letter (A, B, C, or D)
- "explanation": a brief explanation of why the answer is correct

Return a JSON array of objects. The array must contain exactly 10 objects.
Each object must have keys: question, options, answer, explanation.
"""
    response = client.chat.completions.create(
        messages=[
            {"role": "system", "content": "You are an exam creator. Always output valid JSON."},
            {"role": "user", "content": prompt}
        ],
        model="llama-3.3-70b-versatile",
        temperature=0.4,
        response_format={"type": "json_object"},
    )
    try:
        raw = json.loads(response.choices[0].message.content)
        if isinstance(raw, dict):
            if "mcqs" in raw:
                mcqs = raw["mcqs"]
            elif "questions" in raw:
                mcqs = raw["questions"]
            else:
                for v in raw.values():
                    if isinstance(v, list):
                        mcqs = v
                        break
                else:
                    mcqs = []
        elif isinstance(raw, list):
            mcqs = raw
        else:
            mcqs = []
        validated = []
        for q in mcqs:
            if isinstance(q, dict) and all(k in q for k in ("question", "options", "answer", "explanation")):
                validated.append(q)
        return validated
    except Exception as e:
        print(f"⚠️ MCQs parsing error: {e}")
        return []

# --------------------------------------------------------------
# BUILD EXCEL – GEMINI STYLE (3 SHEETS)
# --------------------------------------------------------------

def build_excel(vocab_list, mcqs):
    wb = openpyxl.Workbook()

    # ---------- Sheet 1: Vocabulary ----------
    ws = wb.active
    ws.title = "Daily Star Vocabulary"

    navy = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    light_navy = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    even = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    ws.merge_cells("A1:J1")
    ws["A1"] = "THE DAILY STAR - DAILY VOCABULARY BANK (BCS & JOB PREP SPECIAL)"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = navy
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = "Curated from Today's Headlines, Editorials & Reports | Includes Meanings, Parts of Speech, Bengali Translations & Exam Examples"
    ws["A2"].font = Font(size=10, italic=True, color="FFFFFF")
    ws["A2"].fill = light_navy
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["SL No", "Word / Idiom", "Part of Speech", "Difficulty Level",
               "Bengali Meaning (বাংলা অর্থ)", "English Definition",
               "Synonyms", "Antonyms", "Contextual Example (Daily Star)", "Category / Topic"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = navy
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for idx, item in enumerate(vocab_list, 5):
        ws.row_dimensions[idx].height = 26
        row_data = [
            idx - 4,
            safe_excel_value(item.get("word", "")),
            safe_excel_value(item.get("pos", "")),
            safe_excel_value(item.get("level", "")),
            safe_excel_value(item.get("bengali", "")),
            safe_excel_value(item.get("definition", "")),
            safe_excel_value(item.get("synonyms", "")),
            safe_excel_value(item.get("antonyms", "")),
            safe_excel_value(item.get("example", "")),
            safe_excel_value(item.get("category", ""))
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col)
            cell.value = val
            cell.border = border
            cell.fill = even if idx % 2 == 0 else white
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col in [1, 3, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if col == 4:
                level = str(val).strip().capitalize()
                if level == "Basic":
                    cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                    cell.font = Font(bold=True, color="2E7D32")
                elif level == "Intermediate":
                    cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
                    cell.font = Font(bold=True, color="F57F17")
                elif level == "Advanced":
                    cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                    cell.font = Font(bold=True, color="C62828")

    widths = [8, 20, 15, 16, 28, 35, 32, 30, 45, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A4:J{len(vocab_list)+4}"
    ws.freeze_panes = "B5"

    # ---------- Sheet 2: Summary ----------
    ws2 = wb.create_sheet("BCS & Job Prep Summary")
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "DAILY STAR VOCABULARY ANALYSIS - BCS & JOB EXAM FOCUS"
    ws2["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws2["A1"].fill = navy
    ws2["A1"].alignment = Alignment(horizontal="center")

    level_counts = {"Basic": 0, "Intermediate": 0, "Advanced": 0}
    for item in vocab_list:
        lvl = item.get("level", "Basic")
        level_counts[lvl] = level_counts.get(lvl, 0) + 1

    ws2["A3"] = "VOCABULARY BREAKDOWN BY LEVEL"
    ws2["A3"].font = Font(bold=True, size=12)
    ws2["A4"] = "Difficulty Level"
    ws2["B4"] = "Word Count"
    ws2["C4"] = "% of Total"
    row = 5
    for lvl in ["Basic", "Intermediate", "Advanced"]:
        ws2[f"A{row}"] = lvl
        ws2[f"B{row}"] = level_counts.get(lvl, 0)
        ws2[f"C{row}"] = level_counts.get(lvl, 0) / len(vocab_list) if vocab_list else 0
        ws2[f"C{row}"].number_format = "0.0%"
        row += 1
    ws2[f"A{row}"] = "Total Vocabulary"
    ws2[f"B{row}"] = f"=SUM(B5:B7)"
    ws2[f"C{row}"] = "=SUM(C5:C7)"

    cat_counts = {}
    for item in vocab_list:
        cat = item.get("category", "Uncategorized")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
    ws2["E3"] = "VOCABULARY DISTRIBUTION BY NEWSPAPER SECTOR"
    ws2["E3"].font = Font(bold=True, size=12)
    ws2["E4"] = "Sector / Category"
    ws2["F4"] = "Word Count"
    ws2["G4"] = "Target Job Exams"
    row = 5
    category_order = ["Politics & Governance", "Economy & Finance", "Law & Human Rights",
                      "Public Health & Env", "Crime & Legal Affairs", "Social & Policy Issues"]
    for cat in category_order:
        count = cat_counts.get(cat, 0)
        if count > 0:
            ws2[f"E{row}"] = cat
            ws2[f"F{row}"] = count
            if "Politics" in cat:
                ws2[f"G{row}"] = "BCS Admin, Judiciary, Secretariat"
            elif "Economy" in cat or "Finance" in cat:
                ws2[f"G{row}"] = "Bangladesh Bank, Commercial Banks"
            elif "Law" in cat or "Rights" in cat:
                ws2[f"G{row}"] = "Judicial Service, Assistant Judge, Police"
            elif "Health" in cat or "Env" in cat:
                ws2[f"G{row}"] = "Medical Officer, BCS General Cadre"
            elif "Crime" in cat:
                ws2[f"G{row}"] = "BCS Police, ACC"
            else:
                ws2[f"G{row}"] = "BCS Written, General Knowledge"
            row += 1

    ws2.merge_cells("A10:G10")
    ws2["A10"] = "BANGLADESHI JOB EXAM PREPARATION GUIDELINES (BCS, BANK, NTRCA, PRIMARY)"
    ws2["A10"].font = Font(bold=True, size=12)
    guidelines = [
        "1. BCS Preliminary Strategy: Daily Star editorials contain high-frequency synonyms/antonyms asked directly in BCS Preliminary (20 Marks English). Focus on Advanced words.",
        "2. BCS Written Strategy: Words like 'Rupture', 'Reimagine', 'Transboundary', 'Exacerbate' enhance the standard of translation and passage writing in BCS Written.",
        "3. Bank Recruitment Exams: Bank exams heavily emphasize vocabulary in context, reading comprehension, and fill-in-the-blanks. Learn synonyms and precise nuances.",
        "4. Revision Technique: Practice making sentences using 5 new Daily Star vocabulary words daily to retain Bengali meaning and appropriate usage."
    ]
    for i, text in enumerate(guidelines, start=11):
        ws2.cell(row=i, column=1, value=text).alignment = Alignment(wrap_text=True)

    # ---------- Sheet 3: Practice Test ----------
    if mcqs and isinstance(mcqs, list) and all(isinstance(q, dict) for q in mcqs):
        ws3 = wb.create_sheet("BCS & Bank Practice Set")
        ws3.merge_cells("A1:G1")
        ws3["A1"] = "DAILY STAR VOCABULARY PRACTICE SET (JOB EXAM MODEL TEST)"
        ws3["A1"].font = Font(size=14, bold=True)
        ws3["A1"].alignment = Alignment(horizontal="center")

        ws3.merge_cells("A2:G2")
        ws3["A2"] = "Test your vocabulary knowledge for upcoming BCS Preliminary, Combined Bank Officers, and Primary Teacher exams."
        ws3["A2"].font = Font(size=10, italic=True)
        ws3["A2"].alignment = Alignment(horizontal="center")

        headers_mcq = ["Q. No", "Question / Contextual Sentence", "Option A", "Option B", "Option C", "Option D", "Correct Answer & Explanation"]
        for col, h in enumerate(headers_mcq, 1):
            cell = ws3.cell(row=4, column=col)
            cell.value = h
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for i, q in enumerate(mcqs[:10], 5):
            ws3.cell(row=i, column=1, value=i-4).alignment = Alignment(horizontal="center")
            ws3.cell(row=i, column=2, value=safe_excel_value(q.get("question", ""))).alignment = Alignment(wrap_text=True)
            opts = q.get("options", [])
            for j, opt in enumerate(opts, 3):
                ws3.cell(row=i, column=j, value=safe_excel_value(opt))
            ws3.cell(row=i, column=7, value=safe_excel_value(q.get("answer", "") + " — " + q.get("explanation", ""))).alignment = Alignment(wrap_text=True)

        ws3.column_dimensions["A"].width = 6
        ws3.column_dimensions["B"].width = 40
        ws3.column_dimensions["C"].width = 20
        ws3.column_dimensions["D"].width = 20
        ws3.column_dimensions["E"].width = 20
        ws3.column_dimensions["F"].width = 20
        ws3.column_dimensions["G"].width = 40
        ws3.freeze_panes = "A5"
    else:
        ws3 = wb.create_sheet("BCS & Bank Practice Set")
        ws3.merge_cells("A1:F1")
        ws3["A1"] = "Practice Test – Not Generated (API issue)"
        ws3["A1"].font = Font(size=14, bold=True)
        ws3["A3"] = "No MCQs were generated due to a temporary issue. Try again tomorrow."

    wb.save(EXCEL_FILE)
    return EXCEL_FILE

# --------------------------------------------------------------
# BUILD HTML – MODERN WITH NAVIGATION
# --------------------------------------------------------------

def build_html(vocab_list, mcqs, summary, date_str):
    html_path = DATA_DIR / f"Vocabulary_{date_str}.html"

    # Pre‑process summary
    summary_br = summary.replace('\n', '<br>')

    # --- Navigation links ---
    # Gather all existing daily HTML files (excluding index)
    all_dates = []
    for f in DATA_DIR.glob("Vocabulary_*.html"):
        d = f.stem.replace("Vocabulary_", "")
        all_dates.append(d)
    all_dates = sorted(all_dates)

    # Find current date index
    try:
        idx = all_dates.index(date_str)
    except ValueError:
        idx = -1

    prev_link = ""
    next_link = ""
    if idx > 0:
        prev_date = all_dates[idx - 1]
        prev_link = f'<a href="Vocabulary_{prev_date}.html" class="nav-link">← {prev_date}</a>'
    if idx < len(all_dates) - 1:
        next_date = all_dates[idx + 1]
        next_link = f'<a href="Vocabulary_{next_date}.html" class="nav-link">{next_date} →</a>'

    nav_html = f"""
    <div class="nav-bar">
        {prev_link}
        <span class="nav-spacer">|</span>
        <a href="index.html" class="nav-link">📋 Archive</a>
        <span class="nav-spacer">|</span>
        {next_link}
    </div>
    """

    # --- Group vocabulary by category ---
    categories = {}
    for item in vocab_list:
        cat = item.get('category', 'Miscellaneous')
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(item)

    sorted_cats = sorted(categories.items())

    # Build category sections
    category_html = ""
    for cat, words in sorted_cats:
        category_html += f"""
        <div class="category-section">
            <div class="category-header">
                <h2>{cat}</h2>
                <span class="word-count">{len(words)} words</span>
            </div>
            <div class="table-wrapper">
                <table>
                    <thead>
                        <tr>
                            <th>Word</th>
                            <th>Bengali</th>
                            <th>POS</th>
                            <th>Synonyms</th>
                            <th>Antonyms</th>
                            <th>Example</th>
                        </tr>
                    </thead>
                    <tbody>
        """
        for w in words:
            category_html += f"""
                        <tr>
                            <td class="word-cell"><strong>{w.get('word', '')}</strong></td>
                            <td class="bengali-cell">{w.get('bengali', '')}</td>
                            <td>{w.get('pos', '')}</td>
                            <td>{w.get('synonyms', '')}</td>
                            <td>{w.get('antonyms', '')}</td>
                            <td class="example-cell">{w.get('example', '')}</td>
                        </tr>
            """
        category_html += """
                    </tbody>
                </table>
            </div>
        </div>
        """

    # Build MCQs
    mcq_html = ""
    if mcqs:
        mcq_html = """
        <div class="mcq-section">
            <div class="section-header">
                <h2>📝 Practice Test</h2>
                <span class="badge">10 Questions</span>
            </div>
        """
        for i, q in enumerate(mcqs[:10], 1):
            options_html = "".join(f'<li class="option">{opt}</li>' for opt in q.get('options', []))
            mcq_html += f"""
            <div class="mcq-card">
                <div class="mcq-question">
                    <span class="q-number">{i}.</span>
                    <span>{q['question']}</span>
                </div>
                <ul class="options-list">
                    {options_html}
                </ul>
                <div class="mcq-answer">
                    <strong>✅ Answer:</strong> {q['answer']} – {q['explanation']}
                </div>
            </div>
            """
        mcq_html += """
        </div>
        """

    # Full HTML
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Daily Vocabulary – {date_str}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Noto Sans Bengali', sans-serif;
            background: #f0f4f8;
            color: #1e293b;
            padding: 20px;
            line-height: 1.5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
        }}
        .hero {{
            background: linear-gradient(135deg, #0f2b4b, #1a4a7a);
            color: #ffffff;
            border-radius: 16px;
            padding: 40px 30px;
            margin-bottom: 30px;
            box-shadow: 0 8px 30px rgba(15, 43, 75, 0.25);
        }}
        .hero h1 {{
            font-size: 28px;
            font-weight: 700;
            letter-spacing: -0.5px;
        }}
        .hero .date {{
            font-size: 16px;
            opacity: 0.8;
            margin-top: 6px;
        }}
        .hero .stats {{
            display: flex;
            gap: 24px;
            margin-top: 16px;
            flex-wrap: wrap;
        }}
        .hero .stat-item {{
            background: rgba(255,255,255,0.12);
            padding: 6px 16px;
            border-radius: 40px;
            font-size: 14px;
            font-weight: 500;
        }}
        .hero .stat-item span {{
            font-weight: 700;
        }}
        .nav-bar {{
            margin-top: 16px;
            display: flex;
            gap: 12px;
            align-items: center;
            flex-wrap: wrap;
            font-size: 14px;
        }}
        .nav-link {{
            color: #fff;
            text-decoration: none;
            background: rgba(255,255,255,0.15);
            padding: 4px 14px;
            border-radius: 40px;
            transition: 0.2s;
        }}
        .nav-link:hover {{
            background: rgba(255,255,255,0.25);
        }}
        .nav-spacer {{
            opacity: 0.4;
        }}
        .summary-card {{
            background: #ffffff;
            border-radius: 12px;
            padding: 24px 28px;
            margin-bottom: 30px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            border-left: 5px solid #1a4a7a;
        }}
        .summary-card h3 {{
            font-size: 16px;
            color: #1a4a7a;
            margin-bottom: 10px;
        }}
        .summary-card p {{
            font-size: 15px;
            color: #334155;
            line-height: 1.7;
            white-space: pre-wrap;
        }}
        .category-section {{
            background: #ffffff;
            border-radius: 12px;
            padding: 20px 24px 24px;
            margin-bottom: 24px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        }}
        .category-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 2px solid #e9edf2;
            padding-bottom: 12px;
            margin-bottom: 16px;
        }}
        .category-header h2 {{
            font-size: 18px;
            font-weight: 600;
            color: #0f2b4b;
        }}
        .word-count {{
            font-size: 13px;
            background: #e9edf2;
            padding: 2px 14px;
            border-radius: 40px;
            color: #475569;
            font-weight: 500;
        }}
        .table-wrapper {{
            overflow-x: auto;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
        }}
        thead th {{
            background: #f1f5f9;
            color: #1e293b;
            font-weight: 600;
            padding: 10px 12px;
            text-align: left;
            border-bottom: 2px solid #d1d9e6;
            font-size: 13px;
            text-transform: uppercase;
            letter-spacing: 0.3px;
        }}
        tbody td {{
            padding: 10px 12px;
            border-bottom: 1px solid #e9edf2;
            vertical-align: top;
        }}
        tbody tr:hover {{
            background: #f8fafc;
        }}
        .word-cell {{
            font-weight: 600;
            color: #0f2b4b;
        }}
        .bengali-cell {{
            color: #1e293b;
        }}
        .example-cell {{
            font-style: italic;
            color: #475569;
            font-size: 13px;
        }}
        .mcq-section {{
            background: #ffffff;
            border-radius: 12px;
            padding: 20px 24px 24px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        }}
        .section-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 2px solid #e9edf2;
            padding-bottom: 12px;
            margin-bottom: 20px;
        }}
        .section-header h2 {{
            font-size: 18px;
            font-weight: 600;
            color: #0f2b4b;
        }}
        .badge {{
            font-size: 13px;
            background: #1a4a7a;
            color: #fff;
            padding: 2px 14px;
            border-radius: 40px;
            font-weight: 500;
        }}
        .mcq-card {{
            background: #f8fafc;
            border-radius: 10px;
            padding: 16px 20px;
            margin-bottom: 14px;
            border-left: 4px solid #1a4a7a;
        }}
        .mcq-question {{
            font-weight: 500;
            margin-bottom: 8px;
        }}
        .q-number {{
            color: #1a4a7a;
            font-weight: 700;
            margin-right: 6px;
        }}
        .options-list {{
            list-style: none;
            padding-left: 24px;
            margin-bottom: 8px;
        }}
        .option {{
            font-size: 14px;
            color: #334155;
            padding: 2px 0;
        }}
        .mcq-answer {{
            font-size: 14px;
            color: #0f2b4b;
            padding-top: 6px;
            border-top: 1px dashed #d1d9e6;
        }}
        .footer {{
            text-align: center;
            margin-top: 30px;
            font-size: 13px;
            color: #94a3b8;
            padding: 20px 0 10px;
            border-top: 1px solid #e2e8f0;
        }}
        @media (max-width: 700px) {{
            body {{ padding: 12px; }}
            .hero {{ padding: 24px 18px; }}
            .hero h1 {{ font-size: 22px; }}
            .hero .stats {{ gap: 12px; }}
            .hero .stat-item {{ font-size: 12px; padding: 4px 12px; }}
            .category-section {{ padding: 14px 14px 18px; }}
            .category-header h2 {{ font-size: 16px; }}
            table {{ font-size: 12px; }}
            thead th, tbody td {{ padding: 6px 8px; }}
            .mcq-card {{ padding: 12px 14px; }}
            .options-list {{ padding-left: 16px; }}
        }}
        @media print {{
            body {{ background: #fff; padding: 0; }}
            .hero {{ box-shadow: none; }}
            .category-section, .summary-card, .mcq-section {{ box-shadow: none; border: 1px solid #ddd; }}
        }}
    </style>
</head>
<body>
<div class="container">

    <!-- Hero -->
    <div class="hero">
        <h1>📘 Daily Vocabulary Bank</h1>
        <div class="date">{date_str}</div>
        <div class="stats">
            <div class="stat-item">📚 <span>{len(vocab_list)}</span> words</div>
            <div class="stat-item">📂 <span>{len(categories)}</span> categories</div>
            <div class="stat-item">📝 <span>10</span> MCQs</div>
        </div>
        {nav_html}
    </div>

    <!-- Summary -->
    <div class="summary-card">
        <h3>📰 Today's Summary</h3>
        <p>{summary_br}</p>
    </div>

    <!-- Vocabulary by Category -->
    {category_html}

    <!-- MCQs -->
    {mcq_html}

    <!-- Footer -->
    <div class="footer">
        Generated automatically • Daily Star Vocabulary Bank • For BCS & Bank Exams
    </div>

</div>
</body>
</html>
"""
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    return html_path

# --------------------------------------------------------------
# BUILD INDEX PAGE (ARCHIVE)
# --------------------------------------------------------------

def build_index_html(date_str):
    """Generate a master index page listing all available daily HTML files."""
    # Get all existing Vocabulary_*.html files
    all_files = sorted(DATA_DIR.glob("Vocabulary_*.html"), reverse=True)
    # Build list of links
    links_html = ""
    for f in all_files:
        file_date = f.stem.replace("Vocabulary_", "")
        is_today = (file_date == date_str)
        active_class = ' class="today"' if is_today else ''
        links_html += f'<li{active_class}><a href="{f.name}">{file_date}</a></li>\n'

    index_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Daily Vocabulary Archive</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: #f0f4f8;
            padding: 30px 20px;
            color: #1e293b;
        }}
        .container {{
            max-width: 800px;
            margin: 0 auto;
            background: #fff;
            border-radius: 16px;
            padding: 40px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.06);
        }}
        h1 {{
            font-size: 28px;
            color: #0f2b4b;
            margin-bottom: 6px;
        }}
        .subtitle {{
            color: #64748b;
            font-size: 16px;
            margin-bottom: 24px;
        }}
        ul {{
            list-style: none;
            padding: 0;
        }}
        li {{
            padding: 10px 14px;
            border-bottom: 1px solid #e9edf2;
        }}
        li:last-child {{
            border-bottom: none;
        }}
        li a {{
            text-decoration: none;
            color: #1a4a7a;
            font-weight: 500;
            display: block;
        }}
        li a:hover {{
            color: #0f2b4b;
            text-decoration: underline;
        }}
        .today {{
            background: #e9edf2;
            border-radius: 6px;
            font-weight: 600;
        }}
        .today a {{
            color: #0f2b4b;
        }}
        .footer {{
            margin-top: 30px;
            font-size: 13px;
            color: #94a3b8;
            border-top: 1px solid #e2e8f0;
            padding-top: 16px;
            text-align: center;
        }}
        @media (max-width: 600px) {{
            .container {{ padding: 20px; }}
        }}
    </style>
</head>
<body>
<div class="container">
    <h1>📚 Daily Vocabulary Archive</h1>
    <div class="subtitle">BCS & Bank Exam Preparation</div>
    <ul>
        {links_html}
    </ul>
    <div class="footer">
        Generated automatically • Updated daily
    </div>
</div>
</body>
</html>
"""
    index_path = DATA_DIR / "index.html"
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(index_content)
    return index_path

# --------------------------------------------------------------
# MAIN JOB
# --------------------------------------------------------------

async def run_daily_job():
    print("🚀 Starting daily job...")
    client = groq.Groq(api_key=GROQ_API_KEY)

    print("📰 Fetching headlines...")
    headlines = fetch_headlines()
    print(f"   Found {len(headlines)} headlines.")

    past = load_history()
    print("🧠 Generating vocabulary and summary...")
    vocab, summary = generate_vocab_and_summary(client, headlines, past)
    print(f"   Generated {len(vocab)} words.")

    print("❓ Generating MCQs...")
    mcqs = generate_mcqs(client, vocab)
    print(f"   Generated {len(mcqs)} MCQs.")

    print("📊 Building Excel file...")
    try:
        build_excel(vocab, mcqs)
        print(f"   Excel saved to {EXCEL_FILE}")
    except Exception as e:
        print(f"   ❌ Error building Excel: {e}")

    date_str = datetime.now(TIMEZONE).strftime("%Y-%m-%d")
    print("📄 Building HTML file...")
    try:
        html_path = build_html(vocab, mcqs, summary, date_str)
        print(f"   HTML saved to {html_path}")
    except Exception as e:
        print(f"   ❌ Error building HTML: {e}")

    # Build/update the index page
    print("📋 Building archive index...")
    try:
        index_path = build_index_html(date_str)
        print(f"   Index saved to {index_path}")
    except Exception as e:
        print(f"   ❌ Error building index: {e}")

    save_history(past + [w["word"] for w in vocab])
    print("💾 History updated.")

    # ---------- Telegram Sending ----------
    if TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID:
        print("📤 Sending files via Telegram...")
        try:
            bot = Bot(token=TELEGRAM_BOT_TOKEN)

            if EXCEL_FILE.exists():
                with open(EXCEL_FILE, "rb") as f:
                    await bot.send_document(chat_id=TELEGRAM_CHAT_ID, document=f, caption="📊 Daily Vocabulary Bank (Excel)")
                print("   ✅ Excel sent.")
            else:
                print("   ⚠️ Excel file not found – skipping.")

            if html_path.exists():
                with open(html_path, "rb") as f:
                    await bot.send_document(chat_id=TELEGRAM_CHAT_ID, document=f, caption="📄 Daily Vocabulary Bank (HTML – open on any device)")
                print("   ✅ HTML sent.")
            else:
                print("   ⚠️ HTML file not found – skipping.")

            if not summary or len(summary.strip()) < 10:
                summary = "আজকের সংক্ষিপ্ত সারাংশ তৈরি করা সম্ভব হয়নি। তবে ভোকাবুলারি ফাইলগুলো দেখুন।"
            await bot.send_message(chat_id=TELEGRAM_CHAT_ID, text=f"📰 **Today's Summary**\n\n{summary}", parse_mode="Markdown")
            print("   ✅ Summary sent.")

        except Exception as e:
            print(f"   ❌ Failed to send Telegram messages: {e}")
    else:
        print("⚠️ Telegram credentials missing. Skipping send.")

    print("✅ Job done – Excel, HTML and index generated.")

if __name__ == "__main__":
    asyncio.run(run_daily_job())
